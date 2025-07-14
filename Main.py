import re
import fitz
import docx
import xml.etree.ElementTree as ET
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from difflib import SequenceMatcher
import spacy

# Load NLP model
nlp = spacy.load("en_core_web_sm")

# --- Document extraction functions ---
def extract_text_from_pdf(file):
    text = ""
    doc = fitz.open(stream=file.read(), filetype="pdf")
    for page in doc:
        text += page.get_text()
    return text

def extract_text_from_docx(file):
    doc = docx.Document(file)
    return "\n".join([p.text for p in doc.paragraphs])

def extract_text_from_excel(file):
    wb = load_workbook(file)
    text = ""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    text += f"{cell.value}\n"
    return text

def extract_text_from_xml(file):
    tree = ET.parse(file)
    root = tree.getroot()
    return ET.tostring(root, encoding="unicode")

def extract_parameters(text):
    doc = nlp(text)
    parameters = {}
    for ent in doc.ents:
        if ent.label_ in {"QUANTITY", "CARDINAL", "PERCENT", "MONEY", "ORDINAL", "DATE", "TIME"}:
            parameters[ent.text] = ent.label_
    return parameters

def regex_fallback(text):
    pattern = r"([\w\s]+?)\s*[:\-]\s*([\w\d\.\%\s]+)"
    matches = re.findall(pattern, text)
    return {k.strip(): v.strip() for k, v in matches if k and v}

def is_similar(a, b, threshold=0.8):
    return SequenceMatcher(None, str(a), str(b)).ratio() >= threshold

def process_file(file):
    name = file.name.lower()
    if name.endswith(".pdf"):
        text = extract_text_from_pdf(file)
    elif name.endswith(".docx"):
        text = extract_text_from_docx(file)
    elif name.endswith(".xlsx"):
        text = extract_text_from_excel(file)
    elif name.endswith(".xml"):
        text = extract_text_from_xml(file)
    else:
        return {}, f"Unsupported file type: {name}"
    params = extract_parameters(text)
    if not params:
        params = regex_fallback(text)
    return params, None

def compare_documents(params_list):
    all_keys = set()
    for params in params_list:
        all_keys.update(params.keys())

    data = []
    for key in sorted(all_keys):
        row = {"Parameter": key}
        values = []
        for idx, params in enumerate(params_list, 1):
            val = params.get(key, "❌ Missing")
            row[f"Doc {idx}"] = val
            values.append(val)
        base_val = str(values[0])
        sim_scores = [SequenceMatcher(None, base_val, str(v)).ratio() for v in values]
        confidence = round(sum(sim_scores)/len(sim_scores), 2)
        row["Confidence"] = confidence
        row["Match"] = confidence >= 0.8
        data.append(row)

    return pd.DataFrame(data)

# ---- CUSTOM CSS & THEME ----
st.set_page_config(
    page_title="Stellantis | Parameter Comparison Tool",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
:root {
    --primary-color: #004c97;
    --secondary-color: #0073e6;
    --accent-color: #00bfff;
}
h1, h2, h3, .stButton>button {
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
}
.stButton>button {
    background-color: var(--primary-color);
    color: white;
    border-radius: 8px;
    padding: 8px 18px;
    transition: background-color 0.3s ease;
}
.stButton>button:hover {
    background-color: var(--secondary-color);
    cursor: pointer;
}
.stProgress > div > div {
    background-color: var(--accent-color) !important;
}
[data-testid="stFileUploadDropzone"] {
    border: 3px dashed var(--primary-color);
    border-radius: 12px;
    padding: 20px;
}
</style>
""", unsafe_allow_html=True)

# --- Dark mode toggle ---
if 'dark_mode' not in st.session_state:
    st.session_state.dark_mode = False

def toggle_dark_mode():
    st.session_state.dark_mode = not st.session_state.dark_mode

st.sidebar.title("⚙️ Settings")
st.sidebar.checkbox("Enable Dark Mode", value=st.session_state.dark_mode, on_change=toggle_dark_mode)

if st.session_state.dark_mode:
    st.markdown("""
    <style>
    body, .css-18e3th9 {
        background-color: #121212;
        color: #eeeeee;
    }
    .css-1d391kg {
        background-color: #1e1e1e;
    }
    </style>
    """, unsafe_allow_html=True)

st.title("📊 Stellantis Automated Parameter Comparison Tool")
st.markdown(
    """
    Upload two or more documents to extract structured parameters and compare them side-by-side.
    Supported formats: **PDF, DOCX, XLSX, XML**.
    """)

uploaded_files = st.file_uploader(
    "Upload documents for comparison (min 2 files):",
    type=["pdf", "docx", "xlsx", "xml"],
    accept_multiple_files=True
)

if uploaded_files and len(uploaded_files) >= 2:
    with st.spinner("Analyzing files, please wait... ⏳"):
        params_list = []
        failed = []
        for f in uploaded_files:
            p, err = process_file(f)
            if err:
                failed.append(f.name)
            else:
                params_list.append(p)

    if failed:
        st.error(f"⚠️ Could not process files: {', '.join(failed)}")

    if len(params_list) >= 2:
        comparison_df = compare_documents(params_list)

        def style_row(row):
            if not row.Match:
                return ['background-color: #ffcccc']*len(row)
            return ['']*len(row)

        # Sidebar filter options
        st.sidebar.markdown("### 🔎 Filter Comparison Results")
        filter_option = st.sidebar.radio("Show parameters:", options=["All", "Only Mismatches"])
        search_text = st.sidebar.text_input("Search parameters")

        df_to_show = comparison_df.copy()
        if filter_option == "Only Mismatches":
            df_to_show = df_to_show[df_to_show["Match"] == False]
        if search_text:
            df_to_show = df_to_show[df_to_show["Parameter"].str.contains(search_text, case=False, na=False)]

        # Fixed: Confidence color bar with two colors only
        styled_df = df_to_show.style.apply(style_row, axis=1).format({
            "Confidence": "{:.0%}"
        }).bar(subset=["Confidence"], color=['#f28c8c', '#85e085'])  # red for low, green for high

        st.markdown("### Comparison Results")
        st.dataframe(styled_df, use_container_width=True, height=450)

        # Download buttons
        csv_data = df_to_show.to_csv(index=False).encode('utf-8')
        excel_data = df_to_show.to_excel(index=False, engine='openpyxl')

        st.download_button("📥 Download CSV", csv_data, "stellantis_comparison.csv", "text/csv")
else:
    st.info("Please upload at least two valid documents to start comparison.")
